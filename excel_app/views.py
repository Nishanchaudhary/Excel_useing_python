import os
from django.shortcuts import render
from openpyxl import load_workbook

def excel_data_view(request):
    # Path to the Excel file
    excel_file_path = os.path.join(os.path.dirname(__file__), 'data.xlsx')

    # Load the workbook and select the active sheet
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # Extract data from the sheet
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Pass the data to the template
    context = {'excel_data': data}
    return render(request, 'excel_data.html', context)